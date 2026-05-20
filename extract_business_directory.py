import streamlit as st
import pandas as pd
from io import BytesIO
from PIL import Image
import json

# Correct standard Pydantic import for structured schemas
from pydantic import BaseModel

# Native Google GenAI SDK imports
from google import genai
from google.genai import types

# Initialize the Gemini Client securely via Streamlit Secrets
client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])

# Define the expected JSON output structures using standard Pydantic models
class BusinessRecord(BaseModel):
    name_korean: str
    name_english: str
    street_number: str      # Cleanly isolated street/building number
    street_name: str        # Cleanly isolated street name/suite/route
    city: str
    state: str
    zip_code: str
    phone_number: str
    website_or_notes: str

class DirectoryPage(BaseModel):
    category_name: str
    businesses: list[BusinessRecord]

# Streamlit UI Configuration
st.set_page_config(page_title="Korean Business Directory Extractor", layout="wide")
st.title("📋 Korean Business Directory AI Extractor")
st.write("Upload scanned directory pages to instantly extract and generate multi-tab stylized Excel spreadsheets with split address values.")

uploaded_files = st.file_uploader(
    "Drag and drop directory image pages (JPG/PNG)", 
    type=["jpg", "jpeg", "png"], 
    accept_multiple_files=True
)

if uploaded_files and st.button("🚀 Process Pages & Generate Excel"):
    all_tabs_data = {}
    
    with st.spinner("Analyzing pages with Gemini 3 Flash Preview..."):
        for uploaded_file in uploaded_files:
            st.write(f"⚙️ Processing: `{uploaded_file.name}`...")
            
            # Load image cleanly for processing
            image = Image.open(uploaded_file)
            
            prompt = """
            You are an expert data extraction AI. Look at this Korean business directory page.
            Extract EVERY single business listed on this page. 
            Identify the core business category of this page to use as the tab name.
            For each entry, split the names and parse the address information cleanly. 
            Crucially, break down the street information into two distinct parts:
            1. The numeric house/building number goes into 'street_number'.
            2. The name of the street (and any suite/apartment numbers) goes into 'street_name'.
            """
            
            try:
                # Call Gemini 3 Flash Preview with Structured Outputs enforced via Pydantic schema
                response = client.models.generate_content(
                    model='gemini-3-flash-preview',
                    contents=[image, prompt],
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        response_schema=DirectoryPage,
                        temperature=0.1
                    ),
                )
                
                # Parse response safely
                page_json = json.loads(response.text)
                category = page_json.get("category_name", uploaded_file.name.split(".")[0])
                records = page_json.get("businesses", [])
                
                if records:
                    # Convert parsed records directly to a DataFrame
                    df = pd.DataFrame(records)
                    
                    # Standardize columns to match the new presentation layout with split streets
                    df.columns = [
                        "Name (Korean)", 
                        "Name (English)", 
                        "Street Number", 
                        "Street Name", 
                        "City", 
                        "State", 
                        "Zip Code", 
                        "Phone Number", 
                        "Website / Notes"
                    ]
                    all_tabs_data[category] = df
                    
            except Exception as e:
                st.error(f"❌ Failed to process {uploaded_file.name}: {str(e)}")

    if all_tabs_data:
        # Build stylized Excel file directly in memory buffer
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in all_tabs_data.items():
                # Clean sheet name to fit Excel rules (max 31 characters, no invalid symbols)
                clean_sheet_name = sheet_name[:30].replace("[", "").replace("]", "").replace(":", "").replace("*", "").replace("?", "").replace("/", "").replace("\\", "")
                df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
                
                # Access layout engine for granular styling configuration
                workbook = writer.book
                worksheet = writer.sheets[clean_sheet_name]
                worksheet.views.sheetView[0].showGridLines = True
                
                # Dynamic cell styling properties
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                
                # Format headers row perfectly
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Adjust line height bounds
                worksheet.row_dimensions[1].height = 26
                
                # Auto-adjust column bounds to match data contents
                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

        excel_data = output.getvalue()
        
        st.success("🎉 Extraction Complete for all uploaded pages!")
        st.download_button(
            label="📥 Download Structured Excel File",
            data=excel_data,
            file_name="Extracted_Korean_Business_Directory.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Please upload one or more image files to begin.")